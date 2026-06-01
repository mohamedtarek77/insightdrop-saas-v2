"""
Excel Export Service – returns the cleaned + feature-engineered DataFrame
as a multi-sheet .xlsx workbook, entirely in memory.

Sheet 1 – Cleaned Data      (all rows after dedup/null-handling + derived cols)
Sheet 2 – Monthly Summary
Sheet 3 – Top Products
Sheet 4 – Region Summary
Sheet 5 – Category Summary
"""

from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ── Colour palette ────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="0D0D0F")   # near-black
HDR_FONT   = Font(name="Calibri", bold=True, color="C8FA64", size=10)
ALT_FILL   = PatternFill("solid", fgColor="F0F0F4")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
KPI_FILL   = PatternFill("solid", fgColor="F5F5F7")
KPI_FONT   = Font(name="Calibri", bold=True, size=11, color="0D0D0F")
THIN_SIDE  = Side(style="thin", color="CCCCCC")
THIN_BORDER= Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

def _fmt_currency(v: float) -> str:
    if v >= 1_000_000: return f"${v/1_000_000:.2f}M"
    if v >= 1_000:     return f"${v/1_000:.2f}K"
    return f"${v:.2f}"


def _style_header_row(ws, row: int, col_count: int) -> None:
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_rows(ws, start_row: int, end_row: int, col_count: int) -> None:
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else WHITE_FILL
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill   = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, min_w: int = 10, max_w: int = 40) -> None:
    for col_cells in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value else 0) for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
            max(min_w, min(length + 4, max_w))
        )


def _write_df_to_sheet(ws, df: pd.DataFrame, title: str | None = None) -> None:
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(
            name="Calibri", bold=True, size=13, color="0D0D0F"
        )
        ws.row_dimensions[1].height = 22
        start_row = 3

    # Header
    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=str(col))
    _style_header_row(ws, start_row, len(df.columns))
    ws.row_dimensions[start_row].height = 20

    # Data
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    _style_data_rows(ws, start_row + 1, start_row + len(df), len(df.columns))

    # Freeze header
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    _auto_width(ws)


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel_export(
    cleaned_df: pd.DataFrame,
    analytics: dict[str, Any],
) -> bytes:
    """
    Build a styled multi-sheet Excel workbook.
    Returns raw .xlsx bytes.
    """
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Cleaned Data ─────────────────────────────────────────
        export_cols = [
            "order_id", "product_name", "category", "region",
            "order_date", "quantity", "price", "cost",
            "revenue", "profit", "order_month",
        ]
        # Keep only columns that exist (pipeline may vary)
        available = [c for c in export_cols if c in cleaned_df.columns]
        sheet1 = cleaned_df[available].copy()
        sheet1["order_date"] = sheet1["order_date"].dt.strftime("%Y-%m-%d")
        sheet1.to_excel(writer, sheet_name="Cleaned Data", index=False, startrow=2)

        # ── Sheet 2: Monthly Summary ──────────────────────────────────────
        monthly = pd.DataFrame(analytics["charts"]["monthly_sales"])
        if not monthly.empty:
            monthly["avg_order_value"] = (
                monthly["revenue"] / monthly["orders"].replace(0, 1)
            ).round(2)
            monthly.columns = ["Month", "Revenue ($)", "Orders", "Avg Order Value ($)"]
            monthly.to_excel(writer, sheet_name="Monthly Summary", index=False, startrow=2)

        # ── Sheet 3: Top Products ─────────────────────────────────────────
        products = pd.DataFrame(analytics["charts"]["top_products"])
        if not products.empty:
            products.columns = ["Product", "Revenue ($)", "Units Sold"]
            products.insert(0, "Rank", range(1, len(products) + 1))
            products.to_excel(writer, sheet_name="Top Products", index=False, startrow=2)

        # ── Sheet 4: Region Summary ───────────────────────────────────────
        regions = pd.DataFrame(analytics["charts"]["region_sales"])
        if not regions.empty:
            total_rev = analytics["kpis"]["total_revenue"] or 1
            regions["share_pct"] = (regions["revenue"] / total_rev * 100).round(1)
            regions.columns = ["Region", "Revenue ($)", "Orders", "Share (%)"]
            regions.to_excel(writer, sheet_name="Region Summary", index=False, startrow=2)

        # ── Sheet 5: Category Summary ─────────────────────────────────────
        cats = pd.DataFrame(analytics["charts"]["category_sales"])
        if not cats.empty:
            cats["margin_pct"] = (
                cats["profit"] / cats["revenue"].replace(0, 1) * 100
            ).round(1)
            cats.columns = ["Category", "Revenue ($)", "Profit ($)", "Orders", "Margin (%)"]
            cats.to_excel(writer, sheet_name="Category Summary", index=False, startrow=2)

    # ── Post-process: apply openpyxl styling ─────────────────────────────
    buf.seek(0)
    wb = load_workbook(buf)

    sheet_titles = {
        "Cleaned Data":     "📋  Cleaned & Enriched Sales Data",
        "Monthly Summary":  "📅  Monthly Revenue Summary",
        "Top Products":     "🏆  Top Products by Revenue",
        "Region Summary":   "🌍  Sales by Region",
        "Category Summary": "📦  Category Performance",
    }

    for sheet_name, title in sheet_titles.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Title row (row 1)
        ws.cell(row=1, column=1, value=title).font = Font(
            name="Calibri", bold=True, size=13, color="0D0D0F"
        )
        ws.row_dimensions[1].height = 24

        # Find header row (row 3 = startrow 2 + 1 in 1-based openpyxl)
        header_row = 3
        n_cols = ws.max_column
        _style_header_row(ws, header_row, n_cols)
        ws.row_dimensions[header_row].height = 20

        # Data rows
        _style_data_rows(ws, header_row + 1, ws.max_row, n_cols)

        # Freeze panes
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        _auto_width(ws)

        # Tab colour
        ws.sheet_properties.tabColor = "C8FA64"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
